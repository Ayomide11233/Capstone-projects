#!/usr/bin/env python3
"""
NBA Stats Web Scraper
=====================
Scrapes the top 10 NBA players by Points Per Game from
basketball-reference.com and exports the data to a formatted Excel file.

Usage:
    python nba_scraper.py
    python nba_scraper.py --season 2025 --stat pts
    python nba_scraper.py --season 2024 --stat ast --output my_stats.xlsx

Stat options: pts (points), ast (assists), reb (rebounds), blk (blocks), stl (steals)
"""

import argparse
import datetime
import sys
import time

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ── Config ────────────────────────────────────────────────────────────────

BREF_URL = "https://www.basketball-reference.com/leagues/NBA_{year}_per_game.html"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Referer": "https://www.basketball-reference.com/",
    "DNT": "1",
    "Connection": "keep-alive",
}

STAT_SORT_MAP = {
    "pts": "pts_per_g",
    "ast": "ast_per_g",
    "reb": "trb_per_g",
    "blk": "blk_per_g",
    "stl": "stl_per_g",
}

STAT_LABELS = {
    "pts": "Points Per Game",
    "ast": "Assists Per Game",
    "reb": "Rebounds Per Game",
    "blk": "Blocks Per Game",
    "stl": "Steals Per Game",
}

# Excel colour palette
NAVY     = "1A1F5E"
NAVY2    = "2C3480"
LT_GOLD  = "FFF3CC"
DARK_GR  = "2D2D2D"
MED_GR   = "F2F2F2"
WHITE    = "FFFFFF"
GOLD     = "FFD700"
SILVER   = "C0C0C0"
BRONZE   = "CD7F32"


# ── Scraping ──────────────────────────────────────────────────────────────

def fetch_page(url: str, retries: int = 3) -> BeautifulSoup:
    """Download a page and return a BeautifulSoup object."""
    for attempt in range(1, retries + 1):
        try:
            print(f"  Fetching: {url}  (attempt {attempt})")
            resp = requests.get(url, headers=HEADERS, timeout=15)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "lxml")
        except requests.HTTPError as e:
            print(f"  HTTP error: {e}")
            if attempt == retries:
                raise
        except requests.RequestException as e:
            print(f"  Request error: {e}")
            if attempt == retries:
                raise
        time.sleep(2 ** attempt)


def parse_player_row(row) -> dict | None:
    """Extract per-game stats from a single <tr> row."""
    # Skip header or separator rows
    if row.get("class") and "thead" in row.get("class", []):
        return None
    th = row.find("th", {"data-stat": "ranker"})
    if not th or not th.text.strip().isdigit():
        return None

    def cell(stat):
        td = row.find("td", {"data-stat": stat})
        if td is None:
            return ""
        return td.text.strip()

    # Player name (strip * from HOF eligibility marker)
    player_td = row.find("td", {"data-stat": "player"})
    if not player_td:
        return None
    player_name = player_td.text.strip().rstrip("*")

    # Multi-team players have team == "TOT" — keep only that row
    team = cell("team_id")

    def safe_float(val):
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    return {
        "player":  player_name,
        "team":    team,
        "pos":     cell("pos"),
        "age":     cell("age"),
        "gp":      safe_float(cell("g")),
        "gs":      safe_float(cell("gs")),
        "mpg":     safe_float(cell("mp_per_g")),
        "fgp":     safe_float(cell("fg_pct")) * 100,
        "3pp":     safe_float(cell("fg3_pct")) * 100,
        "ftp":     safe_float(cell("ft_pct")) * 100,
        "rpg":     safe_float(cell("trb_per_g")),
        "apg":     safe_float(cell("ast_per_g")),
        "spg":     safe_float(cell("stl_per_g")),
        "bpg":     safe_float(cell("blk_per_g")),
        "ppg":     safe_float(cell("pts_per_g")),
        "per":     safe_float(cell("per")) if cell("per") else 0.0,
    }


def scrape_top10(season_end_year: int, stat: str = "pts") -> list[dict]:
    """
    Scrape basketball-reference.com per-game stats table.
    Returns the top-10 players sorted by the chosen stat (descending).
    Deduplicates multi-team players (keeps 'TOT' row).
    """
    url = BREF_URL.format(year=season_end_year)
    soup = fetch_page(url)

    table = soup.find("table", {"id": "per_game_stats"})
    if table is None:
        raise RuntimeError("Could not find the per_game_stats table on the page.")

    rows = table.find("tbody").find_all("tr")
    players = []
    seen = set()

    for row in rows:
        data = parse_player_row(row)
        if data is None:
            continue
        # Keep only the TOT (total) row for traded players
        name = data["player"]
        if name in seen and data["team"] != "TOT":
            continue
        if name in seen:  # Replace previous entry with TOT
            players = [p for p in players if p["player"] != name]
        seen.add(name)
        players.append(data)

    sort_key = STAT_SORT_MAP.get(stat, "ppg")
    # Map internal key
    key_map = {
        "pts_per_g": "ppg",
        "ast_per_g": "apg",
        "trb_per_g": "rpg",
        "blk_per_g": "bpg",
        "stl_per_g": "spg",
    }
    sort_field = key_map.get(sort_key, "ppg")
    players.sort(key=lambda p: p[sort_field], reverse=True)
    return players[:10]


# ── Excel helpers ─────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11, italic=False) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_thin  = Side(style="thin",   color="CCCCCC")
_thick = Side(style="medium", color=NAVY)

def _border(thick_bottom=False) -> Border:
    b = _thick if thick_bottom else _thin
    return Border(left=_thin, right=_thin, top=_thin, bottom=b)


# ── Excel export ──────────────────────────────────────────────────────────

def build_excel(players: list[dict], season_end_year: int,
                stat: str, output_path: str) -> None:
    """Write a polished Excel workbook with the top-10 player stats."""
    wb = openpyxl.Workbook()
    stat_label = STAT_LABELS.get(stat, "Points Per Game")
    season_str = f"{season_end_year - 1}–{str(season_end_year)[-2:]}"

    # ── Sheet 1: Main stats table ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Top 10 Players"

    # Title
    ws.merge_cells("A1:O1")
    ws["A1"].value     = f"NBA {season_str} Season  —  Top 10 Players by {stat_label}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=15, color=WHITE)
    ws["A1"].fill      = _fill(NAVY)
    ws["A1"].alignment = _align()
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("A2:O2")
    ws["A2"].value = (f"Source: basketball-reference.com  |  "
                      f"Generated: {datetime.date.today().strftime('%B %d, %Y')}  |  "
                      f"Regular Season")
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="888888")
    ws["A2"].alignment = _align()
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # Column meta
    COL_LETTERS = list("ABCDEFGHIJKLMNO")
    COL_HEADS   = ["Rank","Player","Team","Pos","GP","MPG","PPG","RPG",
                   "APG","SPG","BPG","FG%","3P%","FT%","AGE"]
    COL_DESC    = ["#","Name","Club","Position","Games","Min/G","Pts/G","Reb/G",
                   "Ast/G","Stl/G","Blk/G","FG%","3PT%","FT%","Age"]
    COL_WIDTHS  = [6,26,7,6,6,7,8,7,7,7,7,7,7,7,7]

    # Header rows
    for col_l, head, desc, width in zip(COL_LETTERS, COL_HEADS, COL_DESC, COL_WIDTHS):
        c4 = ws[f"{col_l}4"]
        c4.value     = head
        c4.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        c4.fill      = _fill(NAVY)
        c4.alignment = _align()
        c4.border    = _border(thick_bottom=True)

        c5 = ws[f"{col_l}5"]
        c5.value     = desc
        c5.font      = Font(name="Arial", size=8, italic=True, color=WHITE)
        c5.fill      = _fill(NAVY2)
        c5.alignment = _align()
        c5.border    = _border()

        ws.column_dimensions[col_l].width = width

    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 15

    medal_bg = {"1": (GOLD,"7A5800"), "2": (SILVER,"404040"), "3": (BRONZE,"5C3000")}

    for idx, player in enumerate(players):
        row = 6 + idx
        rank = idx + 1
        bg = _fill(LT_GOLD if idx % 2 == 0 else WHITE)

        vals = [
            rank, player["player"], player["team"], player["pos"],
            int(player["gp"]), round(player["mpg"], 1), round(player["ppg"], 1),
            round(player["rpg"], 1), round(player["apg"], 1),
            round(player["spg"], 1), round(player["bpg"], 1),
            round(player["fgp"], 1), round(player["3pp"], 1),
            round(player["ftp"], 1), player["age"],
        ]

        for col_l, val in zip(COL_LETTERS, vals):
            c = ws[f"{col_l}{row}"]
            c.value     = val
            c.fill      = bg
            c.border    = _border()
            c.alignment = _align("left" if col_l == "B" else "center")

            if col_l == "A":
                rk = str(rank)
                if rk in medal_bg:
                    c.fill = _fill(medal_bg[rk][0])
                    c.font = Font(name="Arial", bold=True, size=12, color=medal_bg[rk][1])
                else:
                    c.font = Font(name="Arial", bold=True, size=11, color=NAVY)
            elif col_l == "B":
                c.font = Font(name="Arial", bold=True, size=11, color=DARK_GR)
            elif col_l == "G":
                c.font = Font(name="Arial", bold=True, size=11, color=NAVY)
                c.fill = _fill(LT_GOLD)
            else:
                c.font = _font(size=10)

        ws.row_dimensions[row].height = 24

    # Averages row
    avg_row = 17
    ws.merge_cells(f"A{avg_row}:D{avg_row}")
    ws[f"A{avg_row}"].value     = "AVERAGES — Top 10"
    ws[f"A{avg_row}"].font      = Font(name="Arial", bold=True, size=9, color=WHITE)
    ws[f"A{avg_row}"].fill      = _fill(DARK_GR)
    ws[f"A{avg_row}"].alignment = _align()
    ws[f"A{avg_row}"].border    = _border()

    for col_l in ["E","F","G","H","I","J","K","L","M","N","O"]:
        c = ws[f"{col_l}{avg_row}"]
        c.value  = f"=AVERAGE({col_l}6:{col_l}15)"
        c.font   = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill   = _fill(DARK_GR)
        c.alignment = _align()
        c.border = _border()
        c.number_format = "0.0"

    ws.row_dimensions[avg_row].height = 22
    ws.freeze_panes = "C6"

    # ── Sheet 2: Bar chart data ───────────────────────────────────────────
    ws2 = wb.create_sheet("Charts")
    chart_heads = ["Player", "PPG", "RPG", "APG"]
    for j, h in enumerate(chart_heads, 1):
        c = ws2.cell(1, j, h)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = _fill(NAVY)
        c.alignment = _align()

    for i, p in enumerate(players):
        r = i + 2
        ws2.cell(r, 1, p["player"])
        ws2.cell(r, 2, round(p["ppg"], 1))
        ws2.cell(r, 3, round(p["rpg"], 1))
        ws2.cell(r, 4, round(p["apg"], 1))
        ws2.row_dimensions[r].height = 18

    chart = BarChart()
    chart.type     = "bar"
    chart.grouping = "clustered"
    chart.title    = f"Top 10 — PPG / RPG / APG  ({season_str})"
    chart.y_axis.title = "Per Game"
    chart.style    = 10
    chart.width    = 30
    chart.height   = 18

    cats = Reference(ws2, min_col=1, min_row=2, max_row=11)
    for col_i, lbl in [(2,"PPG"),(3,"RPG"),(4,"APG")]:
        data = Reference(ws2, min_col=col_i, min_row=1, max_row=11)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "F1")

    ws2.column_dimensions["A"].width = 26
    for cl in ["B","C","D"]:
        ws2.column_dimensions[cl].width = 8

    # ── Sheet 3: Shooting splits ──────────────────────────────────────────
    ws3 = wb.create_sheet("Shooting Splits")

    ws3.merge_cells("A1:F1")
    ws3["A1"].value     = f"Shooting Splits — {season_str} NBA Regular Season"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws3["A1"].fill      = _fill(NAVY)
    ws3["A1"].alignment = _align()
    ws3.row_dimensions[1].height = 30

    split_heads = ["Player","Team","FG%","3P%","FT%","TS% (approx)"]
    for j, h in enumerate(split_heads, 1):
        c = ws3.cell(2, j, h)
        c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill      = _fill(NAVY2)
        c.alignment = _align()

    for i, p in enumerate(players):
        r = i + 3
        bg = _fill(MED_GR if i % 2 == 0 else WHITE)
        ts = round(p["fgp"] * 0.4 + p["3pp"] * 0.1 + p["ftp"] * 0.5, 1)
        row_vals = [p["player"], p["team"], round(p["fgp"],1),
                    round(p["3pp"],1), round(p["ftp"],1), ts]
        haligns = ["left","center","center","center","center","center"]
        for j, (val, ha) in enumerate(zip(row_vals, haligns), 1):
            c = ws3.cell(r, j, val)
            c.fill      = bg
            c.border    = _border()
            c.alignment = _align(ha)
            c.number_format = "0.0"
            c.font = (Font(name="Arial", bold=True, size=10) if j == 1
                      else _font(size=10))

    for j, w in enumerate([26,8,9,9,9,18], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    wb.save(output_path)
    print(f"\n  Excel saved: {output_path}")


# ── CLI ───────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="NBA stats scraper → Excel")
    p.add_argument("--season", type=int, default=datetime.date.today().year,
                   help="Season end year, e.g. 2025 for 2024-25 (default: current year)")
    p.add_argument("--stat", choices=["pts","ast","reb","blk","stl"], default="pts",
                   help="Stat to rank players by (default: pts)")
    p.add_argument("--output", type=str, default="nba_top10_stats.xlsx",
                   help="Output file path (default: nba_top10_stats.xlsx)")
    return p.parse_args()


def main():
    args = parse_args()
    print(f"\nNBA Stats Scraper")
    print(f"  Season : {args.season - 1}-{str(args.season)[-2:]}")
    print(f"  Ranking: {STAT_LABELS[args.stat]}")
    print(f"  Output : {args.output}\n")

    print("Step 1/2 — Scraping basketball-reference.com ...")
    try:
        players = scrape_top10(args.season, args.stat)
    except Exception as e:
        print(f"\nScrape failed: {e}")
        print("Note: basketball-reference.com may be blocking automated requests.")
        print("Try adding a delay, using a VPN, or running from a different network.\n")
        sys.exit(1)

    print(f"\n  Top 10 by {STAT_LABELS[args.stat]}:\n")
    for i, p in enumerate(players, 1):
        print(f"  {i:>2}. {p['player']:<28} {p['team']}  "
              f"PPG:{p['ppg']:>5.1f}  RPG:{p['rpg']:>5.1f}  APG:{p['apg']:>5.1f}")

    print("\nStep 2/2 — Building Excel workbook ...")
    build_excel(players, args.season, args.stat, args.output)
    print("Done!\n")


if __name__ == "__main__":
    main()